"""
Excel Handler Module for Credit Risk Monitoring System
Creates Excel templates and handles data import/export.
"""

import pandas as pd
import numpy as np
from datetime import datetime, date
from typing import Dict, List, Optional, Any
from pathlib import Path
import sys

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from config.config import EXCEL_DIR, DemoConfig, RiskParameters
from src.database import DatabaseManager

# Try to import openpyxl for Excel styling
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available, basic Excel functionality only")


class ExcelTemplateGenerator:
    """
    Generates Excel templates for data import and manages data export.
    """

    def __init__(self, output_dir: Path = None):
        """
        Initialize the Excel handler.

        Args:
            output_dir: Directory for Excel files (default: data/excel_templates)
        """
        self.output_dir = output_dir or EXCEL_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _create_styled_workbook(self) -> 'Workbook':
        """Create a styled workbook with consistent formatting."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required for styled templates")

        wb = Workbook()
        return wb

    def _apply_header_style(self, ws, row: int = 1):
        """Apply header styling to a worksheet row."""
        if not OPENPYXL_AVAILABLE:
            return

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def generate_customer_template(self) -> Path:
        """
        Generate Excel template for customer data import.

        Returns:
            Path to generated template
        """
        output_path = self.output_dir / 'vorlage_kunden.xlsx'

        # Sample data
        sample_data = pd.DataFrame({
            'name': ['Beispiel GmbH', 'Muster AG', 'Test KG'],
            'branche': ['IT_Technologie', 'Maschinenbau', 'Handel'],
            'kreditrating': ['BBB', 'A', 'BB'],
            'gruendungsjahr': [2010, 1995, 2005],
            'bonitaetsindex': [72.5, 85.0, 58.3],
            'region': ['Bayern', 'Baden-Württemberg', 'Hessen'],
            'risiko_klasse': ['mittel', 'niedrig', 'mittel'],
            'kunden_segment': ['sme', 'corporate', 'sme'],
            'umsatz': [5000000, 150000000, 8500000],
            'mitarbeiteranzahl': [45, 850, 65],
            'eigenkapitalquote': [35.5, 42.0, 28.3]
        })

        # Instructions data
        instructions = pd.DataFrame({
            'Spalte': list(sample_data.columns),
            'Beschreibung': [
                'Firmenname des Kunden',
                f'Branche - erlaubt: {", ".join(DemoConfig.INDUSTRIES[:5])}...',
                f'Kreditrating (AAA bis D)',
                'Gründungsjahr',
                'Bonitätsindex (0-100)',
                f'Region - erlaubt: {", ".join(DemoConfig.REGIONS[:5])}...',
                'Risikoklasse: niedrig, mittel, hoch, sehr_hoch',
                'Segment: retail, sme, corporate',
                'Jahresumsatz in EUR',
                'Anzahl Mitarbeiter',
                'Eigenkapitalquote in %'
            ],
            'Pflichtfeld': ['Ja', 'Ja', 'Ja', 'Nein', 'Nein', 'Ja', 'Ja', 'Ja', 'Nein', 'Nein', 'Nein'],
            'Beispiel': [str(v) for v in sample_data.iloc[0].values]
        })

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Instructions sheet
            instructions.to_excel(writer, sheet_name='Anleitung', index=False)

            # Data template sheet
            sample_data.to_excel(writer, sheet_name='Kundendaten', index=False)

            # Reference data sheet
            ref_data = pd.DataFrame({
                'Branchen': DemoConfig.INDUSTRIES + [''] * (16 - len(DemoConfig.INDUSTRIES)),
                'Regionen': DemoConfig.REGIONS,
                'Ratings': ['AAA', 'AA', 'A', 'BBB', 'BB', 'B', 'CCC', 'CC', 'C', 'D', '', '', '', '', '', ''],
                'Risikoklassen': ['niedrig', 'mittel', 'hoch', 'sehr_hoch', '', '', '', '', '', '', '', '', '', '', '', ''],
                'Segmente': ['retail', 'sme', 'corporate', '', '', '', '', '', '', '', '', '', '', '', '', '']
            })
            ref_data.to_excel(writer, sheet_name='Referenzdaten', index=False)

        print(f"Customer template created: {output_path}")
        return output_path

    def generate_contract_template(self) -> Path:
        """
        Generate Excel template for contract data import.

        Returns:
            Path to generated template
        """
        output_path = self.output_dir / 'vorlage_vertraege.xlsx'

        sample_data = pd.DataFrame({
            'kunden_id': [1, 2, 3],
            'produkt_typ': ['Darlehen', 'Kreditlinie', 'Hypothek'],
            'vertragsdatum': ['2024-01-15', '2024-03-20', '2023-11-01'],
            'laufzeit_monate': [60, 24, 240],
            'zinssatz': [0.0525, 0.0875, 0.0385],
            'waehrung': ['EUR', 'EUR', 'EUR'],
            'kreditlimit': [500000, 1000000, 2500000],
            'ausgenutztes_limit': [500000, 650000, 2500000],
            'restschuld': [425000, 650000, 2350000],
            'sicherheiten_wert': [300000, 200000, 3000000],
            'sicherheiten_typ': ['Buergschaft', 'Warenlager', 'Immobilie'],
            'tilgungsart': ['annuitaet', 'endfaellig', 'annuitaet'],
            'zweckbindung': ['Betriebsmittel', 'Expansion', 'Immobilienerwerb']
        })

        instructions = pd.DataFrame({
            'Spalte': list(sample_data.columns),
            'Beschreibung': [
                'ID des Kunden aus Kundendaten',
                f'Produkttyp: {", ".join(DemoConfig.PRODUCT_TYPES)}',
                'Vertragsdatum (YYYY-MM-DD)',
                'Laufzeit in Monaten',
                'Zinssatz als Dezimalzahl (z.B. 0.05 für 5%)',
                'Währungscode (Standard: EUR)',
                'Maximales Kreditlimit',
                'Aktuell ausgenutztes Limit',
                'Aktuelle Restschuld',
                'Wert der Sicherheiten',
                'Art der Sicherheit',
                'Tilgungsart: annuitaet, endfaellig, linear',
                'Verwendungszweck'
            ],
            'Pflichtfeld': ['Ja', 'Ja', 'Ja', 'Ja', 'Ja', 'Ja', 'Ja', 'Nein', 'Nein', 'Nein', 'Nein', 'Nein', 'Nein'],
            'Beispiel': [str(v) for v in sample_data.iloc[0].values]
        })

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            instructions.to_excel(writer, sheet_name='Anleitung', index=False)
            sample_data.to_excel(writer, sheet_name='Vertragsdaten', index=False)

            ref_data = pd.DataFrame({
                'Produkttypen': DemoConfig.PRODUCT_TYPES + [''] * (10 - len(DemoConfig.PRODUCT_TYPES)),
                'Sicherheiten': ['Immobilie', 'Buergschaft', 'Warenlager', 'Forderungen', 'Finanzielle_Sicherheit', 'Keine', 'Sonstige', '', '', ''],
                'Tilgungsarten': ['annuitaet', 'endfaellig', 'linear', '', '', '', '', '', '', ''],
                'Waehrungen': ['EUR', 'USD', 'GBP', 'CHF', '', '', '', '', '', '']
            })
            ref_data.to_excel(writer, sheet_name='Referenzdaten', index=False)

        print(f"Contract template created: {output_path}")
        return output_path

    def generate_payment_template(self) -> Path:
        """
        Generate Excel template for payment data import.

        Returns:
            Path to generated template
        """
        output_path = self.output_dir / 'vorlage_zahlungen.xlsx'

        sample_data = pd.DataFrame({
            'vertrag_id': [1, 1, 2],
            'faelligkeitsdatum': ['2024-11-01', '2024-12-01', '2024-11-15'],
            'zahlungsdatum': ['2024-11-01', '2024-12-03', ''],
            'soll_betrag': [5000.00, 5000.00, 8500.00],
            'ist_betrag': [5000.00, 5000.00, 0],
            'verspaetung_tage': [0, 2, 0],
            'zahlungsstatus': ['puenktlich', 'puenktlich', 'offen'],
            'zahlungsart': ['Tilgung_und_Zinsen', 'Tilgung_und_Zinsen', 'Zinsen'],
            'kommentar': ['', '', 'Ausstehend']
        })

        instructions = pd.DataFrame({
            'Spalte': list(sample_data.columns),
            'Beschreibung': [
                'ID des Vertrags',
                'Fälligkeitsdatum (YYYY-MM-DD)',
                'Tatsächliches Zahlungsdatum (leer wenn noch nicht gezahlt)',
                'Soll-Betrag der Zahlung',
                'Ist-Betrag der Zahlung',
                'Tage Verspätung (0 wenn pünktlich)',
                'Status: puenktlich, verzoegert, ausfall, offen',
                'Art: Tilgung, Zinsen, Tilgung_und_Zinsen, Gebuehr',
                'Optionaler Kommentar'
            ],
            'Pflichtfeld': ['Ja', 'Ja', 'Nein', 'Ja', 'Nein', 'Nein', 'Ja', 'Nein', 'Nein'],
            'Beispiel': [str(v) for v in sample_data.iloc[0].values]
        })

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            instructions.to_excel(writer, sheet_name='Anleitung', index=False)
            sample_data.to_excel(writer, sheet_name='Zahlungsdaten', index=False)

        print(f"Payment template created: {output_path}")
        return output_path

    def generate_economic_data_template(self) -> Path:
        """
        Generate Excel template for economic data import.

        Returns:
            Path to generated template
        """
        output_path = self.output_dir / 'vorlage_wirtschaftsdaten.xlsx'

        sample_data = pd.DataFrame({
            'datum': ['2024-11-01', '2024-11-01', '2024-10-01'],
            'region': ['Bayern', 'Berlin', 'Bayern'],
            'branche': ['IT_Technologie', '', 'Maschinenbau'],
            'ausfallrate_branche': [0.0125, '', 0.0185],
            'konjunktur_index': [102.5, 98.3, 101.2],
            'arbeitslosenquote': [3.8, 8.2, 4.1],
            'zinsniveau': [0.0425, 0.0425, 0.0400],
            'inflation': [2.8, 2.8, 2.9],
            'bip_wachstum': [1.2, 0.5, 1.5],
            'insolvenzquote': [0.0095, 0.0145, 0.0088],
            'quelle': ['Eurostat', 'Bundesagentur', 'Branchenverband']
        })

        instructions = pd.DataFrame({
            'Spalte': list(sample_data.columns),
            'Beschreibung': [
                'Datum der Erhebung (YYYY-MM-DD)',
                'Region',
                'Branche (leer für regionale Gesamtdaten)',
                'Branchenspezifische Ausfallrate',
                'Konjunkturindex (100 = neutral)',
                'Arbeitslosenquote in %',
                'Zinsniveau (EZB Leitzins)',
                'Inflationsrate in %',
                'BIP-Wachstum in %',
                'Insolvenzquote',
                'Datenquelle'
            ],
            'Pflichtfeld': ['Ja', 'Ja', 'Nein', 'Nein', 'Nein', 'Nein', 'Nein', 'Nein', 'Nein', 'Nein', 'Nein'],
            'Beispiel': [str(v) for v in sample_data.iloc[0].values]
        })

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            instructions.to_excel(writer, sheet_name='Anleitung', index=False)
            sample_data.to_excel(writer, sheet_name='Wirtschaftsdaten', index=False)

        print(f"Economic data template created: {output_path}")
        return output_path

    def generate_all_templates(self) -> Dict[str, Path]:
        """
        Generate all Excel templates.

        Returns:
            Dictionary mapping template names to file paths
        """
        templates = {
            'kunden': self.generate_customer_template(),
            'vertraege': self.generate_contract_template(),
            'zahlungen': self.generate_payment_template(),
            'wirtschaftsdaten': self.generate_economic_data_template()
        }

        print(f"\nAll templates generated in: {self.output_dir}")
        return templates


class ExcelDataImporter:
    """
    Imports data from Excel files into the database.
    """

    def __init__(self, db: DatabaseManager):
        """
        Initialize the data importer.

        Args:
            db: DatabaseManager instance
        """
        self.db = db

    def validate_customer_data(self, df: pd.DataFrame) -> List[str]:
        """
        Validate customer data before import.

        Args:
            df: DataFrame with customer data

        Returns:
            List of validation errors (empty if valid)
        """
        errors = []

        required_cols = ['name', 'branche', 'kreditrating', 'region', 'risiko_klasse', 'kunden_segment']
        for col in required_cols:
            if col not in df.columns:
                errors.append(f"Pflichtfeld '{col}' fehlt")
            elif df[col].isna().any():
                errors.append(f"Pflichtfeld '{col}' enthält leere Werte")

        # Validate values
        valid_ratings = ['AAA', 'AA', 'A', 'BBB', 'BB', 'B', 'CCC', 'CC', 'C', 'D']
        if 'kreditrating' in df.columns:
            invalid = df[~df['kreditrating'].isin(valid_ratings)]['kreditrating'].unique()
            if len(invalid) > 0:
                errors.append(f"Ungültige Ratings: {invalid}")

        valid_segments = ['retail', 'sme', 'corporate']
        if 'kunden_segment' in df.columns:
            invalid = df[~df['kunden_segment'].isin(valid_segments)]['kunden_segment'].unique()
            if len(invalid) > 0:
                errors.append(f"Ungültige Segmente: {invalid}")

        return errors

    def import_customers(self, file_path: Path, sheet_name: str = 'Kundendaten') -> int:
        """
        Import customer data from Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet with data

        Returns:
            Number of imported records
        """
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Validate
        errors = self.validate_customer_data(df)
        if errors:
            raise ValueError(f"Validation errors: {'; '.join(errors)}")

        # Import
        records = df.to_dict('records')
        count = self.db.execute_insert_many('kunden', records)

        print(f"Imported {count} customer records")
        return count

    def import_contracts(self, file_path: Path, sheet_name: str = 'Vertragsdaten') -> int:
        """
        Import contract data from Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet with data

        Returns:
            Number of imported records
        """
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Add calculated fields if missing
        if 'pd_wert' not in df.columns:
            df['pd_wert'] = 0.02  # Default PD
        if 'lgd_wert' not in df.columns:
            df['lgd_wert'] = 0.45  # Default LGD
        if 'ead_wert' not in df.columns:
            df['ead_wert'] = df.get('ausgenutztes_limit', df.get('kreditlimit', 0))
        if 'vertrag_status' not in df.columns:
            df['vertrag_status'] = 'aktiv'

        records = df.to_dict('records')
        count = self.db.execute_insert_many('kredit_vertraege', records)

        print(f"Imported {count} contract records")
        return count

    def import_payments(self, file_path: Path, sheet_name: str = 'Zahlungsdaten') -> int:
        """
        Import payment data from Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet with data

        Returns:
            Number of imported records
        """
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Handle missing values
        df['zahlungsdatum'] = df['zahlungsdatum'].replace('', None)
        df['ist_betrag'] = df['ist_betrag'].fillna(0)
        df['verspaetung_tage'] = df['verspaetung_tage'].fillna(0)

        records = df.to_dict('records')
        count = self.db.execute_insert_many('zahlungen', records)

        print(f"Imported {count} payment records")
        return count

    def import_economic_data(self, file_path: Path, sheet_name: str = 'Wirtschaftsdaten') -> int:
        """
        Import economic data from Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet with data

        Returns:
            Number of imported records
        """
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Replace empty strings with None
        df = df.replace('', None)

        records = df.to_dict('records')

        # Import one by one to handle duplicates
        count = 0
        for record in records:
            try:
                self.db.execute_insert('wirtschaftsdaten', record)
                count += 1
            except Exception as e:
                print(f"Skipping duplicate record: {record.get('datum')} / {record.get('region')}")

        print(f"Imported {count} economic data records")
        return count


class ExcelDataExporter:
    """
    Exports data from database to Excel files.
    """

    def __init__(self, db: DatabaseManager):
        """
        Initialize the data exporter.

        Args:
            db: DatabaseManager instance
        """
        self.db = db

    def export_full_database(self, output_path: Path) -> None:
        """
        Export all database tables to Excel.

        Args:
            output_path: Path for output Excel file
        """
        tables = ['kunden', 'kredit_vertraege', 'zahlungen', 'ausfall_ereignisse',
                  'wirtschaftsdaten', 'risiko_limits', 'rating_historie', 'rueckstellungen']

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for table in tables:
                try:
                    df = self.db.execute_dataframe(f"SELECT * FROM {table}")
                    if not df.empty:
                        df.to_excel(writer, sheet_name=table[:31], index=False)
                        print(f"Exported {table}: {len(df)} records")
                except Exception as e:
                    print(f"Error exporting {table}: {e}")

        print(f"\nFull database exported to: {output_path}")

    def export_portfolio_snapshot(self, output_path: Path) -> None:
        """
        Export current portfolio snapshot for reporting.

        Args:
            output_path: Path for output Excel file
        """
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Active contracts with customer info
            query = """
            SELECT
                v.vertrag_id,
                k.name as kunde,
                k.branche,
                k.region,
                k.kreditrating,
                v.produkt_typ,
                v.kreditlimit,
                v.restschuld,
                v.zinssatz,
                v.vertrag_status,
                v.pd_wert,
                v.lgd_wert
            FROM kredit_vertraege v
            JOIN kunden k ON v.kunden_id = k.kunden_id
            WHERE v.vertrag_status = 'aktiv'
            ORDER BY v.restschuld DESC
            """
            df = self.db.execute_dataframe(query)
            df.to_excel(writer, sheet_name='Aktive_Vertraege', index=False)

            # Summary statistics
            summary_query = """
            SELECT
                COUNT(*) as anzahl,
                SUM(restschuld) as gesamt_exposure,
                AVG(zinssatz) as avg_zinssatz,
                AVG(pd_wert) as avg_pd,
                AVG(lgd_wert) as avg_lgd
            FROM kredit_vertraege
            WHERE vertrag_status = 'aktiv'
            """
            summary_df = self.db.execute_dataframe(summary_query)
            summary_df.to_excel(writer, sheet_name='Zusammenfassung', index=False)

        print(f"Portfolio snapshot exported to: {output_path}")


def generate_sample_excel_files(db: DatabaseManager = None):
    """
    Generate sample Excel files with data from the database.

    Args:
        db: Optional DatabaseManager (uses demo if None)
    """
    from src.database import get_demo_db

    if db is None:
        db = get_demo_db()

    # Generate templates
    template_gen = ExcelTemplateGenerator()
    template_gen.generate_all_templates()

    # Export sample data
    exporter = ExcelDataExporter(db)
    exporter.export_full_database(EXCEL_DIR / 'beispiel_datenbank_export.xlsx')
    exporter.export_portfolio_snapshot(EXCEL_DIR / 'beispiel_portfolio_snapshot.xlsx')


if __name__ == "__main__":
    from src.database import get_demo_db

    print("Generating Excel templates and sample files...")
    db = get_demo_db()
    generate_sample_excel_files(db)
